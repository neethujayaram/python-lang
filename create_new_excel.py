import openpyxl

wb=openpyxl.Workbook()
wb.active
sheet1=wb.create_sheet()
sheet1.title="My New sheet1"
# get_sh=wb.get_sheet_names()
work_sheet=wb.get_sheet_by_name("My New sheet1")

# work_sheet=get_sh[0]

work_sheet['A1'].value="Name"
work_sheet['B1'].value="Mobile No."

for i in range(2,12):
    work_sheet.cell(row=i,column=1).value="test"+str(i)
    work_sheet.cell(row=i,column=2).value=9000+i

wb.save('/home/neethu/Documents/example1234.xlsx')